#!/usr/bin/env python3
"""Karşıyaka Ortaokulu kalıp işi — proje dashboard Excel üretici."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[1] / "Kalip_Proje_Dashboard.xlsx"

# --- Proje sabitleri ---
PROJE_ADI = "Karşıyaka Ortaokulu (24 derslik)"
PROJE_NO = "MEBİZ.73-10-25-01-SU-001-R0"
TOPLAM_M2 = 11773.4
BASLANGIC = datetime(2026, 9, 1)
BITIS = datetime(2027, 4, 15)
FIZIKI_BITIS = datetime(2027, 3, 31)
SURE_AY = 7.5
EKIP_PIK = 6

KALEMLER = [
    ("Perde Kalıbı", 4621.5, 39.3, "Bodrum + üst kat perde"),
    ("Döşeme Alt Kalıbı", 4706.5, 40.0, "5 kat döşeme altı"),
    ("Kolon Kalıbı", 883.2, 7.5, "Bodrum + üst kolon"),
    ("Kiriş Kalıbı", 1441.8, 12.2, "5 kat kiriş yan+alt"),
    ("Radye Yan Kalıbı", 120.4, 1.0, "Temel yan"),
]

BOLUMLER = [
    ("Bodrum Bölümü", 1457.1, 12.4),
    ("Üst Kat Bölümü", 10316.3, 87.6),
]

AYLAR = ["Eyl'26", "Eki'26", "Kas'26", "Ara'26", "Oca'27", "Şub'27", "Mar'27", "Nis'27"]
# Aylık montaj hedefi (m²) — toplam 11773.4
AYLIK_HEDEF = [800, 1100, 1400, 1700, 1900, 2000, 1873, 1000]
# Kategori dağılımı (perde, döşeme, kolon+kiriş+radye) — satır bazlı
AYLIK_PERDE = [420, 580, 720, 880, 980, 720, 650, 171]
AYLIK_DOSEME = [0, 80, 180, 420, 620, 880, 920, 586]
AYLIK_DIGER = [380, 440, 500, 400, 300, 400, 303, 243]

ASAMALAR = ["Hazırlık & Ölçü", "Kalıp Montaj", "Beton & Kontrol", "Söküm & Devretme"]

# Renk paleti (cephe dashboard benzeri koyu tema)
BG_DARK = "1A1A2E"
BG_PANEL = "16213E"
BG_CARD = "0F3460"
ACCENT = "E94560"
ACCENT2 = "F5A623"
ACCENT3 = "4ECDC4"
WHITE = "FFFFFF"
LIGHT = "E8E8E8"
MUTED = "8892A0"

THIN = Side(style="thin", color="334155")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(color=WHITE, bold=False, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)


def apply_dark(ws, max_row=50, max_col=20):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill(BG_DARK)


def merge_title(ws, row, text, col_span=12, size=16, color=ACCENT2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font(color=color, bold=True, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = fill(BG_DARK)


def card_header(ws, row, col, text, width=3):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=9, color=ACCENT2)
    c.fill = fill(BG_CARD)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER


def big_number(ws, row, col, value, suffix="", width=3):
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + width - 1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=True, size=18, color=WHITE)
    c.fill = fill(BG_CARD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if suffix:
        ws.cell(row=row + 2, column=col, value=suffix).font = font(size=8, color=MUTED)
        ws.cell(row=row + 2, column=col).alignment = Alignment(horizontal="center")


def build_parametreler(wb: Workbook):
    ws = wb.create_sheet("_Parametreler", 0)
    ws.sheet_state = "hidden"
    headers = ["Anahtar", "Değer"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)
    params = [
        ("proje_adi", PROJE_ADI),
        ("proje_no", PROJE_NO),
        ("toplam_m2", TOPLAM_M2),
        ("baslangic", BASLANGIC.strftime("%d.%m.%Y")),
        ("bitis", BITIS.strftime("%d.%m.%Y")),
        ("fiziki_bitis", FIZIKI_BITIS.strftime("%d.%m.%Y")),
        ("ekip_pik", EKIP_PIK),
        ("sure_ay", SURE_AY),
    ]
    for i, (k, v) in enumerate(params, 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)


def build_ana_sayfa(wb: Workbook):
    ws = wb.create_sheet("Ana Sayfa")
    apply_dark(ws, 55, 16)
    ws.sheet_view.showGridLines = False

    for col, w in enumerate([2, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    merge_title(ws, 1, f"{TOPLAM_M2:,.0f} m² KALIP İŞİ".replace(",", "."), 12, 18)
    ws.cell(row=2, column=1, value=f"PROJE DASHBOARD — {PROJE_ADI}").font = font(color=MUTED, size=10)
    ws.merge_cells("A2:L2")
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- Üst KPI kartları ---
    card_header(ws, 4, 1, "TOPLAM İLERLEME", 2)
    card_header(ws, 4, 3, "PERDE KALIBI", 2)
    card_header(ws, 4, 5, "DÖŞEME KALIBI", 2)
    card_header(ws, 4, 7, "KOLON + KİRİŞ", 2)
    card_header(ws, 4, 9, "PERSONEL (PİK)", 2)
    card_header(ws, 4, 11, "KALAN GÜN", 2)

    ws.merge_cells("A5:B6")
    ws["A5"] = "='Veri Girişi'!B8"
    ws["A5"].number_format = "0%"
    ws["A5"].font = font(bold=True, size=20, color=ACCENT2)
    ws["A5"].fill = fill(BG_CARD)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A7"] = "genel"
    ws["A7"].font = font(size=8, color=MUTED)
    ws["A7"].alignment = Alignment(horizontal="center")

    big_number(ws, 5, 3, 4621, "m²", width=2)
    big_number(ws, 5, 5, 4707, "m²", width=2)
    big_number(ws, 5, 7, 2325, "m²", width=2)
    big_number(ws, 5, 9, EKIP_PIK, "kişi", width=2)
    big_number(ws, 5, 11, "='Veri Girişi'!B18", "gün", width=2)

    # --- Sol panel: tarihler & istatistik ---
    row = 9
    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value="PROJE ÖZETİ").font = font(bold=True, color=ACCENT2)
    row += 1
    ozet = [
        ("Başlangıç", BASLANGIC.strftime("%d.%m.%Y")),
        ("Bitiş", BITIS.strftime("%d.%m.%Y")),
        ("Fiziki Bitiş", FIZIKI_BITIS.strftime("%d.%m.%Y")),
        ("Toplam Metraj", f"{TOPLAM_M2:,.1f} m²".replace(",", ".")),
        ("Bölüm Sayısı", "2 (Bodrum / Üst Kat)"),
        ("Proje Süresi", f"{SURE_AY} ay"),
        ("Ekip", "2 usta + 3 çırak + 1 mühendis"),
        ("Birim Fiyat (teklif)", "500 TL/m²"),
    ]
    for label, val in ozet:
        ws.cell(row=row, column=1, value=label).font = font(size=9, color=MUTED)
        ws.cell(row=row, column=2, value=val).font = font(size=9, color=WHITE)
        row += 1

    # İş kalemi tablosu (pasta grafik kaynağı)
    row += 1
    ws.cell(row=row, column=1, value="Kalem").font = font(bold=True, color=ACCENT2)
    ws.cell(row=row, column=2, value="m²").font = font(bold=True, color=ACCENT2)
    ws.cell(row=row, column=3, value="Pay %").font = font(bold=True, color=ACCENT2)
    chart_start = row + 1
    for i, (ad, m2, pay, _) in enumerate(KALEMLER):
        r = chart_start + i
        ws.cell(row=r, column=1, value=ad).font = font(size=9)
        ws.cell(row=r, column=2, value=m2).font = font(size=9)
        ws.cell(row=r, column=3, value=pay / 100).number_format = "0.0%"
        ws.cell(row=r, column=3).font = font(size=9)
    chart_end = chart_start + len(KALEMLER) - 1

    # Doughnut — iş kalemi dağılımı
    pie = DoughnutChart()
    pie.title = "İş Kalemleri Dağılımı"
    pie.style = 10
    labels = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
    data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_end)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.holeSize = 55
    pie.width = 11
    pie.height = 9
    ws.add_chart(pie, "D9")

    # --- KPI ilerleme çubukları (sağ üst) ---
    kpi_row = 9
    ws.cell(row=kpi_row, column=12, value="PROJE KPI").font = font(bold=True, color=ACCENT2)
    kpis = [
        ("Hazırlık & Ölçü", "='Veri Girişi'!C4"),
        ("Kalıp Montaj", "='Veri Girişi'!C5"),
        ("Beton & Kontrol", "='Veri Girişi'!C6"),
        ("Söküm & Devretme", "='Veri Girişi'!C7"),
        ("Genel İlerleme", "='Veri Girişi'!C8"),
    ]
    for i, (name, formula) in enumerate(kpis):
        r = kpi_row + 1 + i
        ws.cell(row=r, column=11, value=name).font = font(size=9, color=MUTED)
        ws.cell(row=r, column=12, value=formula).number_format = "0%"
        ws.cell(row=r, column=12).font = font(bold=True, size=11, color=ACCENT3)
        ws.cell(row=r, column=12).fill = fill(BG_CARD)
        ws.cell(row=r, column=12).alignment = Alignment(horizontal="center")

    # --- Aylık montaj hedefi tablosu (grafik kaynağı) ---
    tbl_row = 20
    ws.cell(row=tbl_row, column=1, value="Ay").font = font(bold=True, color=ACCENT2)
    ws.cell(row=tbl_row, column=2, value="Perde").font = font(bold=True, color=ACCENT2)
    ws.cell(row=tbl_row, column=3, value="Döşeme").font = font(bold=True, color=ACCENT2)
    ws.cell(row=tbl_row, column=4, value="Kolon+Kiriş").font = font(bold=True, color=ACCENT2)
    ws.cell(row=tbl_row, column=5, value="Toplam").font = font(bold=True, color=ACCENT2)
    ws.cell(row=tbl_row, column=6, value="Kümülatif %").font = font(bold=True, color=ACCENT2)
    for i, ay in enumerate(AYLAR):
        r = tbl_row + 1 + i
        ws.cell(row=r, column=1, value=ay).font = font(size=9)
        ws.cell(row=r, column=2, value=AYLIK_PERDE[i])
        ws.cell(row=r, column=3, value=AYLIK_DOSEME[i])
        ws.cell(row=r, column=4, value=AYLIK_DIGER[i])
        ws.cell(row=r, column=5, value=AYLIK_HEDEF[i])
        cum = sum(AYLIK_HEDEF[: i + 1]) / TOPLAM_M2
        ws.cell(row=r, column=6, value=cum).number_format = "0%"

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.title = "Aylık Montaj Hedefi (m²)"
    bar.style = 10
    bar.width = 18
    bar.height = 10
    cats = Reference(ws, min_col=1, min_row=tbl_row + 1, max_row=tbl_row + len(AYLAR))
    for col in (2, 3, 4):
        data = Reference(ws, min_col=col, min_row=tbl_row, max_row=tbl_row + len(AYLAR))
        bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "F20")

    # Kümülatif ilerleme çizgi
    line = LineChart()
    line.title = "Kümülatif İlerleme (Plan)"
    line.style = 10
    line.width = 18
    line.height = 10
    line_data = Reference(ws, min_col=6, min_row=tbl_row, max_row=tbl_row + len(AYLAR))
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "F35")

    # --- Bölüm ilerleme kartları ---
    sec_row = 50
    for idx, (bolum, m2, pay) in enumerate(BOLUMLER):
        col = 1 + idx * 6
        ws.merge_cells(start_row=sec_row, start_column=col, end_row=sec_row, end_column=col + 4)
        title_cell = ws.cell(row=sec_row, column=col, value=f"{bolum} — {m2:,.0f} m²".replace(",", "."))
        title_cell.font = font(bold=True, color=ACCENT2, size=11)
        for j, asama in enumerate(ASAMALAR):
            r = sec_row + 2 + j
            ws.cell(row=r, column=col, value=asama).font = font(size=9, color=MUTED)
            formula_col = "B" if idx == 0 else "E"
            pct = ws.cell(row=r, column=col + 1, value=f"='Veri Girişi'!{formula_col}{11 + j}")
            pct.number_format = "0%"
            pct.font = font(bold=True, size=12, color=ACCENT3)
            ws.cell(row=r, column=col + 2, value="DEVAM EDİYOR" if j < 2 else "BEKLEMEDE").font = font(
                size=8, color=MUTED
            )

    # Yönetici notu (alt satır — bölüm kartlarıyla çakışmasın)
    note_row = 58
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 5, end_column=12)
    note = ws.cell(
        row=note_row,
        column=1,
        value=(
            f"Yönetici Notu: {PROJE_ADI} kalıp taşeronluğu {BASLANGIC.strftime('%d.%m.%Y')} — "
            f"{BITIS.strftime('%d.%m.%Y')} arasında planlanmıştır. Toplam {TOPLAM_M2:,.0f} m² kırık ölçü metraj; "
            f"bodrum ve üst kat olmak üzere 2 bölümde ilerlenecektir. Ekip kapasitesi {EKIP_PIK} kişi; "
            f"gerçekçi tamamlanma {FIZIKI_BITIS.strftime('%d.%m.%Y')} hedeflenmektedir. "
            f"Güncel ilerleme 'Veri Girişi' sayfasından güncellenir."
        ).replace(",", "."),
    )
    note.font = font(size=9, color=LIGHT)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    note.fill = fill(BG_CARD)


def build_veri_girisi(wb: Workbook):
    ws = wb.create_sheet("Veri Girişi")
    ws.sheet_view.showGridLines = True

    ws["A1"] = "KALIP PROJE — VERİ GİRİŞİ & İLERLEME TAKİBİ"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:H1")

    ws["A3"] = "Genel Aşama İlerlemesi (%)"
    ws["A3"].font = Font(bold=True)
    for i, a in enumerate(ASAMALAR + ["Genel İlerleme"], 4):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=0 if i < 8 else "=AVERAGE(B4:B7)")
        ws.cell(row=i, column=2).number_format = "0%"
        if i == 8:
            ws.cell(row=i, column=2).font = Font(bold=True)

    ws["A10"] = "Bodrum Bölümü İlerlemesi (%)"
    ws["A10"].font = Font(bold=True)
    for i, a in enumerate(ASAMALAR, 11):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=0)
        ws.cell(row=i, column=2).number_format = "0%"

    ws["D10"] = "Üst Kat Bölümü İlerlemesi (%)"
    ws["D10"].font = Font(bold=True)
    for i, a in enumerate(ASAMALAR, 11):
        ws.cell(row=i, column=4, value=a)
        ws.cell(row=i, column=5, value=0)
        ws.cell(row=i, column=5).number_format = "0%"

    # Kalan gün
    ws["A17"] = "Rapor Tarihi"
    ws["B17"] = "=TODAY()"
    ws["B17"].number_format = "DD.MM.YYYY"
    ws["A18"] = "Kalan Gün"
    ws["B18"] = f"=MAX(0,DATE(2027,4,15)-B17)"

    # Aylık gerçekleşen montaj
    ws["A21"] = "Aylık Gerçekleşen Montaj (m²)"
    ws["A21"].font = Font(bold=True)
    headers = ["Ay", "Plan (m²)", "Gerçekleşen (m²)", "Sapma", "Kümülatif Gerçekleşen", "Kümülatif %"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=22, column=c, value=h).font = Font(bold=True)
    for i, ay in enumerate(AYLAR):
        r = 23 + i
        ws.cell(row=r, column=1, value=ay)
        ws.cell(row=r, column=2, value=AYLIK_HEDEF[i])
        ws.cell(row=r, column=3, value=0)  # kullanıcı girer
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        ws.cell(row=r, column=5, value=f"=SUM($C$23:C{r})")
        ws.cell(row=r, column=6, value=f"=E{r}/{TOPLAM_M2}").number_format = "0.0%"

    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=1, allow_blank=True)
    dv.error = "0 ile 1 arasında girin (örn. 0.25 = %25)"
    ws.add_data_validation(dv)
    for r in range(4, 8):
        dv.add(ws.cell(row=r, column=2))
    for r in range(11, 15):
        dv.add(ws.cell(row=r, column=2))
        dv.add(ws.cell(row=r, column=5))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14


def build_bodrum_takip(wb: Workbook):
    ws = wb.create_sheet("Bodrum Takip")
    ws["A1"] = "BODRUM BÖLÜMÜ — KALIP TAKİP"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    kalemler = [
        ("Radye yan kalıp", 120.4),
        ("Bodrum perde kalıbı", 1152.7),
        ("Bodrum kolon kalıbı", 184.0),
    ]
    headers = ["Kalem", "Metraj (m²)", "Montaj %", "Beton %", "Söküm %", "Tamamlanan (m²)", "Kalan (m²)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)
    for i, (ad, m2) in enumerate(kalemler, 4):
        ws.cell(row=i, column=1, value=ad)
        ws.cell(row=i, column=2, value=m2)
        ws.cell(row=i, column=3, value=0).number_format = "0%"
        ws.cell(row=i, column=4, value=0).number_format = "0%"
        ws.cell(row=i, column=5, value=0).number_format = "0%"
        ws.cell(row=i, column=6, value=f"=B{i}*AVERAGE(C{i}:E{i})")
        ws.cell(row=i, column=7, value=f"=B{i}-F{i}")
    r = 4 + len(kalemler)
    ws.cell(row=r, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").font = Font(bold=True)
    ws.cell(row=r, column=6, value=f"=SUM(F4:F{r-1})").font = Font(bold=True)
    ws.cell(row=r, column=7, value=f"=SUM(G4:G{r-1})").font = Font(bold=True)

    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 16


def build_ust_kat_takip(wb: Workbook):
    ws = wb.create_sheet("Üst Kat Takip")
    ws["A1"] = "ÜST KAT BÖLÜMÜ — KALIP TAKİP"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    kalemler = [
        ("Üst kat perde kalıbı", 3468.8),
        ("Üst kat kolon kalıbı", 699.2),
        ("Döşeme alt kalıbı", 4706.5),
        ("Kiriş kalıbı", 1441.8),
    ]
    headers = ["Kalem", "Metraj (m²)", "Montaj %", "Beton %", "Söküm %", "Tamamlanan (m²)", "Kalan (m²)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)
    for i, (ad, m2) in enumerate(kalemler, 4):
        ws.cell(row=i, column=1, value=ad)
        ws.cell(row=i, column=2, value=m2)
        ws.cell(row=i, column=3, value=0).number_format = "0%"
        ws.cell(row=i, column=4, value=0).number_format = "0%"
        ws.cell(row=i, column=5, value=0).number_format = "0%"
        ws.cell(row=i, column=6, value=f"=B{i}*AVERAGE(C{i}:E{i})")
        ws.cell(row=i, column=7, value=f"=B{i}-F{i}")
    r = 4 + len(kalemler)
    ws.cell(row=r, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").font = Font(bold=True)

    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 18


def build_hakedis_takip(wb: Workbook):
    ws = wb.create_sheet("Hakediş Takip")
    ws["A1"] = "HAKEDİŞ & FİNANS TAKİBİ"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = "Birim fiyat: 500 TL/m² (KDV hariç) · Tevkifat 4/10"

    headers = ["Dönem", "Tamamlanan (m²)", "Matrah (TL)", "KDV %20", "Tevkifat", "Tahsil KDV", "Fatura Tutarı", "Tahsil Edildi"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h).font = Font(bold=True)
    for i in range(1, 9):
        r = 4 + i
        ws.cell(row=r, column=1, value=f"Hakediş {i}")
        ws.cell(row=r, column=2, value=0)
        ws.cell(row=r, column=3, value=f"=B{r}*500")
        ws.cell(row=r, column=4, value=f"=C{r}*0.2")
        ws.cell(row=r, column=5, value=f"=D{r}*0.4")
        ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        ws.cell(row=r, column=7, value=f"=C{r}+F{r}")
        ws.cell(row=r, column=8, value="Hayır")
        for c in range(3, 8):
            ws.cell(row=r, column=c).number_format = '#,##0 "TL"'

    r = 13
    ws.cell(row=r, column=1, value="TOPLAM").font = Font(bold=True)
    for c in range(2, 8):
        ws.cell(row=r, column=c, value=f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}12)").font = Font(bold=True)

    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 16


def build_ekip_puantaj(wb: Workbook):
    ws = wb.create_sheet("Ekip & Puantaj")
    ws["A1"] = "EKİP & PUANTAJ PLANLAMASI"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    ws["A3"] = "Kadro"
    ws["A3"].font = Font(bold=True)
    kadro = [
        ("Kalıp ustası", 2, 120000, "4A"),
        ("Çırak / yardımcı", 3, 60000, "4A"),
        ("Şantiye mühendisi", 1, 70000, "4A"),
        ("Şirket sahibi", 1, "-", "4B Bağ-Kur"),
    ]
    headers = ["Pozisyon", "Adet", "Brüt Aylık (TL)", "SGK"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h).font = Font(bold=True)
    for i, row_data in enumerate(kadro, 5):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=i, column=c, value=val)

    ws["A11"] = "Aylık Personel Planı"
    ws["A11"].font = Font(bold=True)
    ws.cell(row=12, column=1, value="Ay").font = Font(bold=True)
    ws.cell(row=12, column=2, value="Planlanan Ekip").font = Font(bold=True)
    ws.cell(row=12, column=3, value="Gerçekleşen Gün-Kişi").font = Font(bold=True)
    for i, ay in enumerate(AYLAR):
        r = 13 + i
        ws.cell(row=r, column=1, value=ay)
        ws.cell(row=r, column=2, value=EKIP_PIK)
        ws.cell(row=r, column=3, value=0)

    bar = BarChart()
    bar.title = "Aylık Personel Planlaması"
    bar.style = 10
    bar.width = 16
    bar.height = 9
    cats = Reference(ws, min_col=1, min_row=13, max_row=20)
    data = Reference(ws, min_col=2, min_row=12, max_row=20)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "E11")

    for col in "ABCD":
        ws.column_dimensions[col].width = 22


def build_kpi(wb: Workbook):
    ws = wb.create_sheet("KPI")
    ws["A1"] = "KPI & PERFORMANS GÖSTERGELERİ"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    kpis = [
        ("Planlanan m² / gün (ekip)", 68.8, "m²/gün", "Gerçekçi senaryo"),
        ("Toplam metraj", TOPLAM_M2, "m²", "Kırık ölçü"),
        ("Sözleşme süresi", SURE_AY, "ay", "Önerilen"),
        ("Ekip verimliliği hedefi", 85, "%", "Plan/gerçekleşen"),
        ("İSG olay sayısı", 0, "adet", "Hedef: 0"),
        ("Hakediş gecikme", 0, "gün", "Hedef: ≤15"),
        ("Kalite red oranı", 0, "%", "Hedef: <2%"),
        ("Maliyet sapması", 0, "%", "Hedef: ±5%"),
    ]
    headers = ["KPI", "Hedef", "Birim", "Not"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)
    for i, row_data in enumerate(kpis, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=i, column=c, value=val)

    ws["A14"] = "Gerçekleşen KPI (saha girişi)"
    ws["A14"].font = Font(bold=True)
    for i in range(4, 12):
        ws.cell(row=i, column=5, value="Gerçekleşen")
        ws.cell(row=i, column=6, value=0)

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 22


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_parametreler(wb)
    build_veri_girisi(wb)
    build_ana_sayfa(wb)
    build_bodrum_takip(wb)
    build_ust_kat_takip(wb)
    build_hakedis_takip(wb)
    build_ekip_puantaj(wb)
    build_kpi(wb)

    # Ana Sayfa genel ilerleme — A5 formülü build_ana_sayfa içinde

    wb.save(OUT)
    print(f"Oluşturuldu: {OUT}")


if __name__ == "__main__":
    main()
