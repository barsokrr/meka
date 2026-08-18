#!/usr/bin/env python3
"""Karşıyaka kalıp işi — tek sayfa özet tablo Excel."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "Kalip_Ozet_Tablo.xlsx"

HEADER = PatternFill("solid", fgColor="1F4E79")
SECTION = PatternFill("solid", fgColor="D6E4F0")
HIGHLIGHT = PatternFill("solid", fgColor="E2EFDA")
WARN = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROWS = [
    ("PROJE", "Proje adı", "Karşıyaka Ortaokulu (24 derslik)", "", ""),
    ("PROJE", "Proje no", "MEBİZ.73-10-25-01-SU-001-R0", "", ""),
    ("PROJE", "İş tanımı", "Betonarme ahşap kalıp işçiliği (taşeron)", "", ""),
    ("PROJE", "Ölçü yöntemi", "Kırık ölçü (a×b×c)", "", "Tercih"),
    ("PROJE", "Alternatif ölçü", "Düz ölçü", "11.295,7 m²", ""),
    ("TAŞERON", "Unvan", "ABDURRAHMAN BARIŞ ÖKER", "", ""),
    ("TAŞERON", "VKN", "6530560679", "", "Van VD"),
    ("TAŞERON", "NACE", "43.99.05 — Kalıp işleri", "", "Aktif"),
    ("TAŞERON", "Model", "Seçenek A — 4A + Aracı No + tevkifat 4/10", "", ""),
    ("METRAJ", "Radye yan kalıp", 120.4, "m²", "1,0%"),
    ("METRAJ", "Bodrum perde", 1152.7, "m²", "9,8%"),
    ("METRAJ", "Bodrum kolon", 184.0, "m²", "1,6%"),
    ("METRAJ", "Üst kat perde", 3468.8, "m²", "29,5%"),
    ("METRAJ", "Üst kat kolon", 699.2, "m²", "5,9%"),
    ("METRAJ", "Döşeme alt", 4706.5, "m²", "40,0%"),
    ("METRAJ", "Kiriş", 1441.8, "m²", "12,2%"),
    ("METRAJ", "TOPLAM", 11773.4, "m²", "100%"),
    ("FİYAT 450", "Birim fiyat", 450, "TL/m²", "Kanonik"),
    ("FİYAT 450", "Matrah", 5298030, "TL", "KDV hariç"),
    ("FİYAT 450", "Ödenecek", 5933794, "TL", "Tevkifatlı"),
    ("FİYAT 500", "Birim fiyat", 500, "TL/m²", "TEKLİF"),
    ("FİYAT 500", "Matrah", 5886700, "TL", "KDV hariç"),
    ("FİYAT 500", "Ödenecek", 6593104, "TL", "Tevkifatlı"),
    ("FİYAT 550", "Birim fiyat", 550, "TL/m²", "Hedef"),
    ("FİYAT 550", "Matrah", 6475370, "TL", "KDV hariç"),
    ("FİYAT 550", "Ödenecek", 7252414, "TL", "Tevkifatlı"),
    ("REFERANS", "ÇŞB poz 15.180.1002", 917, "TL/m²", "Tam poz"),
    ("REFERANS", "Okul bütçe (tahmini)", "~129,9 M", "TL", "Kalıp ~%4,1"),
    ("EKİP", "Usta", "2 × 120.000", "TL/ay", "4A"),
    ("EKİP", "Çırak", "3 × 60.000", "TL/ay", "4A"),
    ("EKİP", "Mühendis", "1 × 70.000", "TL/ay", "4A"),
    ("EKİP", "Personel pik", 6, "kişi", ""),
    ("MALİYET", "Gider (6 ay)", 3951500, "TL", "Ref."),
    ("MALİYET", "Gider (7,1 ay)", 4628275, "TL", "Gerçekçi"),
    ("KÂR 7.1ay", "450 TL/m² net", 491341, "TL", "Sınırda"),
    ("KÂR 7.1ay", "500 TL/m² net", 873976, "TL", "Yapılabilir"),
    ("KÂR 7.1ay", "550 TL/m² net", 1256612, "TL", "Değer"),
    ("SÜRE", "Gerçekçi", "~7,1", "ay", ""),
    ("SÜRE", "Sözleşme önerisi", "7,5", "ay", ""),
    ("SÜRE", "Plan", "01.09.2026 — 15.04.2027", "", ""),
    ("DURUM", "Aracı No", "Alınmadı", "", "KRİTİK"),
    ("DURUM", "SGK kodu", "Düzeltilmeli", "", "Kalıpçı+MYB"),
    ("DURUM", "Sözleşme", "Taslak", "", "Avukat"),
    ("DURUM", "BF pazarlığı", "Bekliyor", "", "500–550"),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Özet Tablo"

    ws.merge_cells("A1:E1")
    ws["A1"] = "KARŞIYAKA ORTAOKULU — KALIP TAŞERON ÖZET TABLO"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = "MEBİZ.73-10-25-01-SU-001-R0 · 11.773,4 m² kırık ölçü · Taslak (MM/avukat/İSG onayı)"
    ws["A2"].font = Font(size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Bölüm", "Alan", "Değer", "Birim", "Not"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    prev_section = None
    row = 5
    for section, alan, deger, birim, not_ in ROWS:
        if section != prev_section:
            prev_section = section
        for c, val in enumerate([section, alan, deger, birim, not_], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if c == 1:
                cell.fill = SECTION
                cell.font = Font(bold=True, size=9)
            if "TOPLAM" in str(alan) or "TEKLİF" in str(not_):
                cell.fill = HIGHLIGHT
                cell.font = Font(bold=True)
            if section == "DURUM" and "KRİTİK" in str(not_):
                cell.fill = WARN
        row += 1

    # Hızlı karar kutusu
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="HIZLI KARAR").font = Font(bold=True, size=11, color="1F4E79")
    row += 1
    karar = [
        ("Teklif ver", "500 TL/m² → 6.593.104 TL tevkifatlı"),
        ("Minimum", "450 TL/m² → 6 ayda bitirme şart"),
        ("Hedef", "550 TL/m² → ~1,26 M net kâr (7,1 ay)"),
        ("Yasak", "Müteahhit SGK + size fatura (muvazaa)"),
        ("İlk iş", "Aracı No + SGK kalıpçı kodu + MYB"),
    ]
    for label, text in karar:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=text)
        row += 1

    widths = {1: 12, 2: 28, 3: 22, 4: 10, 5: 18}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(OUT)
    print(f"Oluşturuldu: {OUT}")


if __name__ == "__main__":
    main()
