#!/usr/bin/env python3
"""Karşıyaka Ortaokulu kalıp taşeronluğu — işveren fiyat teklifi Excel üretici."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "Isveren_Fiyat_Teklifi_Kalip_Taseronlugu.xlsx"

KALEMLER = [
    ("15.180.1002/A", "Radye temel yan kalıbı yapılması (işçilik)", 120.4),
    ("15.180.1002/B", "Bodrum perde kalıbı yapılması (çift yüz, boşluk düşümlü, işçilik)", 1152.7),
    ("15.180.1002/C", "Bodrum kolon kalıbı yapılması (işçilik)", 184.0),
    ("15.180.1002/D", "Üst kat perde kalıbı yapılması (4 kat aralığı, işçilik)", 3468.8),
    ("15.180.1002/E", "Üst kat kolon kalıbı yapılması (4 kat aralığı, işçilik)", 699.2),
    ("15.180.1002/F", "Döşeme alt kalıbı yapılması (5 kat, işçilik)", 4706.5),
    ("15.180.1002/G", "Kiriş kalıbı yapılması — yan + alt (5 kat, işçilik)", 1441.8),
]

TOPLAM_M2 = sum(k[2] for k in KALEMLER)
TEKLIF_BF = 500  # önerilen teklif birim fiyatı
SENARYOLAR = [450, 500, 550]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ACCENT_FILL = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def money(v: float) -> float:
    return round(v, 2)


def tevkifat_hesap(matrah: float) -> dict:
    kdv = money(matrah * 0.20)
    tevkifat = money(kdv * 0.40)
    tahsil_kdv = money(kdv - tevkifat)
    odenecek = money(matrah + tahsil_kdv)
    return {
        "matrah": money(matrah),
        "kdv": kdv,
        "tevkifat": tevkifat,
        "tahsil_kdv": tahsil_kdv,
        "odenecek": odenecek,
    }


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_table(ws, start_row: int, end_row: int, cols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c >= 5:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_teklif_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Fiyat Teklifi"
    set_col_widths(ws, {"A": 5, "B": 16, "C": 52, "D": 8, "E": 12, "F": 14, "G": 16})

    ws.merge_cells("A1:G1")
    ws["A1"] = "FİYAT TEKLİFİ — KALIP İŞÇİLİĞİ (TAŞERON)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Teklif Veren", "ABDURRAHMAN BARIŞ ÖKER (Gerçek kişi / şahıs işletmesi)"),
        ("Vergi Dairesi / VKN", "Van Vergi Dairesi Müdürlüğü / 6530560679"),
        ("Adres", "Yeni Mah. Çalıbaşı Van İpekyolu No:26/1, Van"),
        ("NACE", "43.99.05 — İnşaatlarda beton işleri (kalıp)"),
        ("", ""),
        ("Proje", "Karşıyaka Ortaokulu (24 derslik)"),
        ("Proje No", "MEBİZ.73-10-25-01-SU-001-R0"),
        ("İş Tanımı", "Betonarme ahşap kalıp işçiliği (taşeron / alt yüklenici)"),
        ("Ölçü Yöntemi", "Kırık ölçü (a × b × c) — plan metraj cetveline göre"),
        ("Teklif Tarihi", date.today().strftime("%d.%m.%Y")),
        ("Geçerlilik", "30 gün"),
        ("Tahmini Süre", "7,5 ay (saha + hava koşulları dahil)"),
    ]
    row = 3
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ["Sıra", "Poz", "İş Kalemi Açıklaması", "Birim", "Miktar (m²)", "Birim Fiyat (TL)", "Tutar (TL)"]
    hdr_row = row
    for c, h in enumerate(headers, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws, hdr_row, len(headers))

    row += 1
    data_start = row
    for i, (poz, aciklama, m2) in enumerate(KALEMLER, 1):
        tutar = money(m2 * TEKLIF_BF)
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=poz)
        ws.cell(row=row, column=3, value=aciklama)
        ws.cell(row=row, column=4, value="m²")
        ws.cell(row=row, column=5, value=m2)
        ws.cell(row=row, column=6, value=TEKLIF_BF)
        ws.cell(row=row, column=7, value=tutar)
        row += 1

    style_table(ws, data_start, row - 1, 7)
    for r in range(data_start, row):
        ws.cell(row=r, column=5).number_format = "#,##0.0"
        ws.cell(row=r, column=6).number_format = '#,##0 "TL"'
        ws.cell(row=r, column=7).number_format = '#,##0.00 "TL"'

    # Toplam satırı
    matrah = money(TOPLAM_M2 * TEKLIF_BF)
    ws.cell(row=row, column=3, value="GENEL TOPLAM").font = Font(bold=True)
    ws.cell(row=row, column=4, value="m²")
    ws.cell(row=row, column=5, value=TOPLAM_M2).font = Font(bold=True)
    ws.cell(row=row, column=6, value=TEKLIF_BF).font = Font(bold=True)
    ws.cell(row=row, column=7, value=matrah).font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = ACCENT_FILL
        ws.cell(row=row, column=c).border = BORDER
    ws.cell(row=row, column=5).number_format = "#,##0.0"
    ws.cell(row=row, column=7).number_format = '#,##0.00 "TL"'
    row += 2

    h = tevkifat_hesap(matrah)
    summary = [
        ("Matrah (KDV hariç)", h["matrah"]),
        ("KDV (%20)", h["kdv"]),
        ("KDV Tevkifatı (4/10)", h["tevkifat"]),
        ("Tahsil edilecek KDV (6/10)", h["tahsil_kdv"]),
        ("TOPLAM ÖDENECEK (matrah + 6/10 KDV)", h["odenecek"]),
    ]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="FİNANSAL ÖZET").font = Font(bold=True, size=11)
    row += 1
    for label, val in summary:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=6, value=val)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6).number_format = '#,##0.00 "TL"'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
        if "TOPLAM" in label:
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = TOTAL_FILL
                ws.cell(row=row, column=c).font = Font(bold=True)
        row += 1

    row += 1
    notes = [
        "Kapsam dahil: Kalıp imalatı, montaj, söküm, işçilik; işçi SGK prim yansıtması; saha iş güvenliği uygulamaları (taşeron personeli).",
        "Kapsam dışı: Kalıp malzemesi, iskele malzemesi, beton, demir, nakliye, vinç/ekipman temini (işveren tarafından sağlanır).",
        "Ödeme: Dönemsel hakediş — çift imzalı metraj tutanağı + e-Fatura (KDV tevkifatı 4/10 uygulanır).",
        "Operasyon: Seçenek A — işçiler taşeron şahıs işletmesi SGK (4A) kaydı + müteahhit dosyasında Aracı No.",
        "Metraj revizyonu: Plan değişikliği / ek metraj durumunda birim fiyat sabit, miktar güncellenir.",
        "Not: Birim fiyat taşeron işçilik bedelidir; ÇŞB 15.180.1002 (2026 Nis. ~917 TL/m²) tam poz referansıdır (malzeme+işçilik+GG+kâr).",
    ]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="TEKLİF ŞARTLARI").font = Font(bold=True, size=11)
    row += 1
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=f"• {note}")
        cell.alignment = Alignment(wrap_text=True)
        row += 1

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=3)
    ws.cell(row=row + 1, column=1, value="Teklif Veren")
    ws.merge_cells(start_row=row + 1, start_column=5, end_row=row + 1, end_column=7)
    ws.cell(row=row + 1, column=5, value="İşveren / Müteahhit")
    ws.cell(row=row + 3, column=1, value="ABDURRAHMAN BARIŞ ÖKER")
    ws.cell(row=row + 3, column=5, value="Unvan: ___________________________")
    ws.cell(row=row + 4, column=5, value="Yetkili: ___________________________")
    ws.cell(row=row + 5, column=5, value="Tarih / İmza: _____________________")


def build_senaryo_sheet(wb: Workbook):
    ws = wb.create_sheet("Birim Fiyat Senaryoları")
    set_col_widths(ws, {"A": 22, "B": 16, "C": 18, "D": 18, "E": 18, "F": 18})

    ws.merge_cells("A1:F1")
    ws["A1"] = "BİRİM FİYAT SENARYO KARŞILAŞTIRMASI"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")

    ws["A3"] = "Toplam metraj (kırık ölçü):"
    ws["B3"] = TOPLAM_M2
    ws["B3"].number_format = "#,##0.0"
    ws["A3"].font = Font(bold=True)

    headers = ["Senaryo", "Birim Fiyat (TL/m²)", "Matrah (KDV hariç)", "KDV %20", "Tevkifat 4/10", "Ödenecek (matrah+6/10 KDV)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, len(headers))

    row = 6
    for bf in SENARYOLAR:
        h = tevkifat_hesap(TOPLAM_M2 * bf)
        label = f"{bf} TL/m²"
        if bf == TEKLIF_BF:
            label += " ← TEKLİF"
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=bf)
        ws.cell(row=row, column=3, value=h["matrah"])
        ws.cell(row=row, column=4, value=h["kdv"])
        ws.cell(row=row, column=5, value=h["tevkifat"])
        ws.cell(row=row, column=6, value=h["odenecek"])
        if bf == TEKLIF_BF:
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = TOTAL_FILL
                ws.cell(row=row, column=c).font = Font(bold=True)
        row += 1

    style_table(ws, 6, row - 1, 6)
    for r in range(6, row):
        for c in range(2, 7):
            ws.cell(row=r, column=c).number_format = '#,##0.00 "TL"'


def build_metraj_sheet(wb: Workbook):
    ws = wb.create_sheet("Metraj Detay")
    set_col_widths(ws, {"A": 5, "B": 16, "C": 48, "D": 14, "E": 12})

    headers = ["Sıra", "Poz", "Kalem", "Metraj (m²)", "Pay (%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    pays = [1.0, 9.8, 1.6, 29.5, 5.9, 40.0, 12.2]
    row = 2
    for i, ((poz, aciklama, m2), pay) in enumerate(zip(KALEMLER, pays), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=poz)
        ws.cell(row=row, column=3, value=aciklama.split(" yapılması")[0])
        ws.cell(row=row, column=4, value=m2)
        ws.cell(row=row, column=5, value=pay)
        row += 1

    ws.cell(row=row, column=3, value="GENEL TOPLAM").font = Font(bold=True)
    ws.cell(row=row, column=4, value=TOPLAM_M2).font = Font(bold=True)
    ws.cell(row=row, column=5, value=100).font = Font(bold=True)
    style_table(ws, 2, row, 5)
    for r in range(2, row + 1):
        ws.cell(row=r, column=4).number_format = "#,##0.0"
        ws.cell(row=r, column=5).number_format = "0.0"


def main():
    wb = Workbook()
    build_teklif_sheet(wb)
    build_senaryo_sheet(wb)
    build_metraj_sheet(wb)
    wb.save(OUT)
    print(f"Oluşturuldu: {OUT}")


if __name__ == "__main__":
    main()
