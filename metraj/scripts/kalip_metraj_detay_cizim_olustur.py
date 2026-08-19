#!/usr/bin/env python3
"""Kalıp metraj detay çizimleri — SVG + HTML + Excel özet."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CIZIM = ROOT / "cizimler"
OUT_HTML = Path(__file__).resolve().parents[2] / "public" / "kalip" / "metraj-cizim.html"
OUT_XLSX = ROOT / "Kalip_Metraj_Detay_Cizim.xlsx"

L, W = 51.18, 20.09  # kat brüt m
CORE = 45.0

KOTLAR = [
    ("Temel", "Radye", "h=0,80 m", 120.4, {"perde": 0, "kolon": 0, "doseme": 0, "kiris": 0}),
    ("Bodrum", "±0,00 altı", "H=4,00 m", 1336.0, {"perde": 1152.7, "kolon": 184.0, "doseme": 0, "kiris": 0}),
    ("±0,00", "Zemin", "H=3,80 m", 2059.2, {"perde": 770.8, "kolon": 174.6, "doseme": 943.4, "kiris": 170.4}),
    ("+4,00", "1. Kat", "H=3,80 m", 2227.7, {"perde": 770.8, "kolon": 174.6, "doseme": 943.4, "kiris": 339.0}),
    ("+8,00", "2. Kat", "H=3,80 m", 2199.7, {"perde": 770.8, "kolon": 174.6, "doseme": 943.4, "kiris": 310.9}),
    ("+12,00", "3. Kat", "H=3,80 m", 2199.7, {"perde": 770.8, "kolon": 174.6, "doseme": 943.4, "kiris": 310.9}),
    ("+16,00", "Çatı", "—", 1253.9, {"perde": 0, "kolon": 0, "doseme": 943.4, "kiris": 310.5}),
]

POZlar = [
    ("A", "Radye temel yan kalıbı", 120.4, "150,5 × 0,80", "1,0%"),
    ("B", "Bodrum perde (çift yüz)", 1152.7, "2×133,8×4,0 − boşluk", "9,8%"),
    ("C", "Bodrum kolon", 184.0, "45,96×4,0×13", "1,6%"),
    ("D", "Üst kat perde ×4", 3468.8, "2×103,1×3,8×4 − boşluk", "29,5%"),
    ("E", "Üst kat kolon ×4", 699.2, "45,96×3,8×13×4", "5,9%"),
    ("F", "Döşeme alt ×5 kat", 4706.5, "(1028,2−düşüm)×5", "40,0%"),
    ("G", "Kiriş yan+alt ×5", 1441.8, "L×(0,80+0,325)", "12,2%"),
]

COLORS = {
    "perde": "#2563eb",
    "kolon": "#dc2626",
    "doseme": "#16a34a",
    "kiris": "#d97706",
    "core": "#94a3b8",
    "radye": "#7c3aed",
    "beam": "#b45309",
    "void": "#f472b6",
}


def svg_plan_typical() -> str:
    scale = 7.5
    pw, ph = L * scale, W * scale
    pad = 40
    cw, ch = pw + pad * 2, ph + pad * 2 + 95
    cx, cy, cw_core, ch_core = pw * 0.55, ph * 0.35, 9 * scale, 5 * scale
    cols = [
        (pw * 0.12, ph * 0.2), (pw * 0.35, ph * 0.2), (pw * 0.75, ph * 0.2),
        (pw * 0.12, ph * 0.75), (pw * 0.35, ph * 0.75), (pw * 0.75, ph * 0.75),
        (pw * 0.5, ph * 0.55), (pw * 0.85, ph * 0.5),
    ]
    col_circles = "".join(
        f'<circle cx="{pad + x}" cy="{pad + y}" r="6" fill="{COLORS["kolon"]}"/>'
        for x, y in cols
    )
    # kiriş hatları şematik
    beams = (
        f'<line x1="{pad+7}" y1="{pad+ph*0.45}" x2="{pad+pw-7}" y2="{pad+ph*0.45}" '
        f'stroke="{COLORS["kiris"]}" stroke-width="3" opacity="0.6"/>'
        f'<line x1="{pad+7}" y1="{pad+ph*0.65}" x2="{pad+pw-7}" y2="{pad+ph*0.65}" '
        f'stroke="{COLORS["kiris"]}" stroke-width="3" opacity="0.6"/>'
    )
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {cw} {ch}" width="100%" style="max-width:560px;background:#fafafa">
  <defs>
    <pattern id="grid" width="20" height="20" patternUnits="userSpaceOnUse">
      <path d="M 20 0 L 0 0 0 20" fill="none" stroke="#e2e8f0" stroke-width="0.5"/>
    </pattern>
    <marker id="arrow" markerWidth="6" markerHeight="6" refX="5" refY="3" orient="auto">
      <path d="M0,0 L6,3 L0,6 Z" fill="#334155"/>
    </marker>
  </defs>
  <text x="{cw/2}" y="22" text-anchor="middle" font-size="13" font-weight="700" fill="#1f4e79">01 — TİPİK KAT PLANI (±0,00 … +12,00)</text>
  <text x="{cw/2}" y="38" text-anchor="middle" font-size="9" fill="#64748b">Brüt {L}×{W} m = 1.028,2 m² · Net döşeme 941,3 m²/kat · Perde CL 103,1 m/kat</text>
  <g transform="translate({pad},{pad})">
    <rect width="{pw}" height="{ph}" fill="url(#grid)"/>
    <rect width="{pw}" height="{ph}" fill="none" stroke="{COLORS['perde']}" stroke-width="14" opacity="0.35"/>
    <rect width="{pw}" height="{ph}" fill="none" stroke="{COLORS['perde']}" stroke-width="2"/>
    <rect x="7" y="7" width="{pw-14}" height="{ph-14}" fill="{COLORS['doseme']}" opacity="0.15"/>
    {beams}
    <rect x="{cx}" y="{cy}" width="{cw_core}" height="{ch_core}" fill="{COLORS['core']}" opacity="0.4" stroke="#64748b" stroke-width="1.5" stroke-dasharray="4,3"/>
    <text x="{cx + cw_core/2}" y="{cy + ch_core/2 - 4}" text-anchor="middle" font-size="8" fill="#475569">MERDİVEN</text>
    <text x="{cx + cw_core/2}" y="{cy + ch_core/2 + 8}" text-anchor="middle" font-size="8" fill="#475569">+ ASANSÖR</text>
    <text x="{cx + cw_core/2}" y="{cy + ch_core/2 + 20}" text-anchor="middle" font-size="7" fill="#64748b">45 m² düşüm</text>
    {col_circles}
    <text x="{pw*0.08}" y="{ph*0.12}" font-size="7" fill="{COLORS['perde']}" font-weight="600">PERDE</text>
    <text x="{pw*0.55}" y="{ph*0.42}" font-size="7" fill="{COLORS['kiris']}">KİRİŞ H1</text>
  </g>
  <line x1="{pad}" y1="{pad+ph+15}" x2="{pad+pw}" y2="{pad+ph+15}" stroke="#334155" marker-end="url(#arrow)" marker-start="url(#arrow)"/>
  <text x="{pad+pw/2}" y="{pad+ph+28}" text-anchor="middle" font-size="10" fill="#334155">{L} m</text>
  <line x1="{pad+pw+15}" y1="{pad}" x2="{pad+pw+15}" y2="{pad+ph}" stroke="#334155" marker-end="url(#arrow)" marker-start="url(#arrow)"/>
  <text x="{pad+pw+28}" y="{pad+ph/2}" text-anchor="middle" font-size="10" fill="#334155" transform="rotate(90 {pad+pw+28} {pad+ph/2})">{W} m</text>
  <g transform="translate({pad}, {pad+ph+45})">
    <rect x="0" y="0" width="12" height="12" fill="{COLORS['perde']}" opacity="0.5"/><text x="16" y="10" font-size="9">Perde 770,8 m²/kat (çift yüz)</text>
    <rect x="170" y="0" width="12" height="12" fill="{COLORS['doseme']}" opacity="0.5"/><text x="186" y="10" font-size="9">Döşeme alt 943,4 m²/kat</text>
    <rect x="0" y="16" width="12" height="12" fill="{COLORS['kolon']}"/><text x="16" y="26" font-size="9">Kolon 174,6 m²/kat · 13 ad</text>
    <rect x="170" y="16" width="12" height="12" fill="{COLORS['kiris']}" opacity="0.6"/><text x="186" y="26" font-size="9">Kiriş 170–339 m²/kat</text>
    <rect x="0" y="32" width="12" height="12" fill="{COLORS['core']}" opacity="0.5"/><text x="16" y="42" font-size="9">Boşluk düşümü 45 m²</text>
  </g>
</svg>'''


def svg_temel_plan() -> str:
    """Temel/bodrum planı — CAD eleman numaraları şematik."""
    pad = 35
    pw, ph = 320, 220
    cw, ch = pw + pad * 2, ph + pad * 2 + 110
    g1, g2, g3 = pad + 80, pad + 200, pad + 320

    def pier(x, y, w, h, label, sub=""):
        sub_t = f'<tspan x="{x + w/2}" dy="10" font-size="6">{sub}</tspan>' if sub else ""
        return (
            f'<rect x="{x}" y="{y}" width="{w}" height="{h}" fill="{COLORS["perde"]}" opacity="0.35" '
            f'stroke="{COLORS["perde"]}" stroke-width="1.2"/>'
            f'<text x="{x + w/2}" y="{y + h/2 + 3}" text-anchor="middle" font-size="7" fill="#1e40af" font-weight="600">'
            f'{label}{sub_t}</text>'
        )

    piers = (
        pier(pad + 20, pad + 30, 18, 55, "P110")
        + pier(g1 - 9, pad + 40, 18, 80, "P109", "30/160")
        + pier(pad + 15, pad + 100, 90, 12, "P122", "180/30")
        + pier(pad + 15, pad + 150, 90, 12, "P124", "180/30")
        + pier(g2 + 30, pad + 80, 110, 14, "P103", "280/35")
        + pier(g1 - 8, pad + 130, 16, 50, "BP17", "30/290")
        + pier(g2 - 8, pad + 130, 16, 55, "BP18", "30/300")
        + pier(g3 - 25, pad + 20, 14, 120, "BP1", "35/575")
    )
    void = (
        f'<rect x="{g2 - 35}" y="{pad + 95}" width="70" height="45" fill="{COLORS["void"]}" opacity="0.25" '
        f'stroke="{COLORS["void"]}" stroke-dasharray="3,2"/>'
        f'<text x="{g2}" y="{pad + 122}" text-anchor="middle" font-size="7" fill="#be185d">BOŞLUK</text>'
        f'<text x="{g2}" y="{pad + 132}" text-anchor="middle" font-size="6" fill="#be185d">100/220</text>'
    )
    radye = (
        f'<rect x="{pad + 5}" y="{pad + 5}" width="{pw - 10}" height="{ph - 10}" fill="{COLORS["radye"]}" '
        f'opacity="0.08" stroke="{COLORS["radye"]}" stroke-width="1.5" stroke-dasharray="6,3"/>'
        f'<text x="{pad + pw/2}" y="{pad + ph - 8}" text-anchor="middle" font-size="8" fill="{COLORS["radye"]}">'
        f'RD01 RADYE · T.Ü.K. −5,10 · h=80 cm</text>'
    )
    kot_markers = (
        f'<circle cx="{g2}" cy="{pad + 60}" r="8" fill="none" stroke="#0284c7" stroke-width="1.5"/>'
        f'<text x="{g2 + 14}" y="{pad + 58}" font-size="7" fill="#0284c7">T.Ü.K. −5,10</text>'
        f'<text x="{g2 - 55}" y="{pad + 118}" font-size="7" fill="#0284c7">B.A.K. −4,50</text>'
    )
    grids = (
        f'<line x1="{g1}" y1="{pad}" x2="{g1}" y2="{pad + ph}" stroke="#cbd5e1" stroke-width="1"/>'
        f'<line x1="{g2}" y1="{pad}" x2="{g2}" y2="{pad + ph}" stroke="#cbd5e1" stroke-width="1"/>'
        f'<line x1="{g3}" y1="{pad}" x2="{g3}" y2="{pad + ph}" stroke="#cbd5e1" stroke-width="1"/>'
        f'<text x="{g1}" y="{pad - 8}" text-anchor="middle" font-size="9" fill="#64748b">1</text>'
        f'<text x="{g2}" y="{pad - 8}" text-anchor="middle" font-size="9" fill="#64748b">2</text>'
        f'<text x="{g3}" y="{pad - 8}" text-anchor="middle" font-size="9" fill="#64748b">3</text>'
    )
    dims = (
        f'<line x1="{pad}" y1="{pad + ph + 12}" x2="{pad + pw}" y2="{pad + ph + 12}" stroke="#0891b2" stroke-width="0.8"/>'
        f'<text x="{pad + pw/2}" y="{pad + ph + 24}" text-anchor="middle" font-size="7" fill="#0891b2">695+360 cm (şema)</text>'
    )
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {cw} {ch}" width="100%" style="max-width:560px;background:#0f172a;color:#fff">
  <rect width="{cw}" height="{ch}" fill="#0f172a"/>
  <text x="{cw/2}" y="22" text-anchor="middle" font-size="13" font-weight="700" fill="#38bdf8">02 — TEMEL / BODRUM PLANI (CAD)</text>
  <text x="{cw/2}" y="38" text-anchor="middle" font-size="9" fill="#94a3b8">Izgara 1–2–3 · Perde + radye · Metraj: bodrum 1.336,7 m²</text>
  {grids}{radye}{piers}{void}{kot_markers}{dims}
  <g transform="translate({pad}, {pad + ph + 40})">
    <text x="0" y="0" font-size="8" fill="#94a3b8">Metraj ilişkisi:</text>
    <text x="0" y="14" font-size="8" fill="#e2e8f0">• Bodrum perde 1.152,7 m² (CL 133,8 m × H 4,0 × 2 yüz)</text>
    <text x="0" y="28" font-size="8" fill="#e2e8f0">• Bodrum kolon 184,0 m² (13 ad · çevre 45,96 m)</text>
    <text x="0" y="42" font-size="8" fill="#e2e8f0">• Radye yan 120,4 m² (çevre 150,5 m × h 0,80)</text>
    <text x="0" y="56" font-size="8" fill="#fbbf24">• BOŞLUK düşümü metrajda düşülmüştür</text>
  </g>
</svg>'''


def svg_cati_plan() -> str:
    """+20,00 kot kalıp planı — CAD."""
    pad = 40
    pw, ph = 270, 272  # 675×680 oran
    cw, ch = pw + pad * 2, ph + pad * 2 + 100
    x0, y0 = pad, pad + 20
    beams = (
        f'<rect x="{x0}" y="{y0}" width="{pw}" height="12" fill="{COLORS["kiris"]}" opacity="0.5"/>'
        f'<rect x="{x0}" y="{y0 + ph - 12}" width="{pw}" height="12" fill="{COLORS["kiris"]}" opacity="0.5"/>'
        f'<rect x="{x0}" y="{y0}" width="12" height="{ph}" fill="{COLORS["kiris"]}" opacity="0.5"/>'
        f'<rect x="{x0 + pw - 12}" y="{y0}" width="12" height="{ph}" fill="{COLORS["kiris"]}" opacity="0.5"/>'
    )
    labels = (
        f'<text x="{x0 + 30}" y="{y0 + 8}" font-size="6" fill="#fff">K0601</text>'
        f'<text x="{x0 + pw - 50}" y="{y0 + 8}" font-size="6" fill="#fff">K0602</text>'
        f'<text x="{x0 + pw/2}" y="{y0 + ph/2}" text-anchor="middle" font-size="10" fill="{COLORS["doseme"]}" font-weight="600">D601 + D602</text>'
        f'<text x="{x0 + pw/2}" y="{y0 + ph/2 + 14}" text-anchor="middle" font-size="8" fill="#64748b">h=20 cm · 650×650 iç</text>'
    )
    grids = (
        f'<line x1="{x0 + pw*0.25}" y1="{y0 - 15}" x2="{x0 + pw*0.25}" y2="{y0 + ph}" stroke="#cbd5e1"/>'
        f'<line x1="{x0 + pw*0.75}" y1="{y0 - 15}" x2="{x0 + pw*0.75}" y2="{y0 + ph}" stroke="#cbd5e1"/>'
        f'<line x1="{x0 - 15}" y1="{y0 + ph*0.3}" x2="{x0 + pw}" y2="{y0 + ph*0.3}" stroke="#cbd5e1"/>'
        f'<line x1="{x0 - 15}" y1="{y0 + ph*0.85}" x2="{x0 + pw}" y2="{y0 + ph*0.85}" stroke="#cbd5e1"/>'
        f'<text x="{x0 + pw*0.25}" y="{y0 - 20}" text-anchor="middle" font-size="9" fill="#64748b">5</text>'
        f'<text x="{x0 + pw*0.75}" y="{y0 - 20}" text-anchor="middle" font-size="9" fill="#64748b">6</text>'
        f'<text x="{x0 - 22}" y="{y0 + ph*0.3}" font-size="9" fill="#64748b">A</text>'
        f'<text x="{x0 - 22}" y="{y0 + ph*0.85}" font-size="9" fill="#64748b">C</text>'
    )
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {cw} {ch}" width="100%" style="max-width:400px;background:#fafafa">
  <text x="{cw/2}" y="18" text-anchor="middle" font-size="12" font-weight="700" fill="#5b21b6">03 — +20,00 KOTU KALIP PLANI</text>
  <text x="{cw/2}" y="32" text-anchor="middle" font-size="8" fill="#64748b">675×680 cm dış · Kesit A-A / B-B · 1/50</text>
  {grids}
  <rect x="{x0 + 12}" y="{y0 + 12}" width="{pw - 24}" height="{ph - 24}" fill="{COLORS['doseme']}" opacity="0.2" stroke="{COLORS['doseme']}" stroke-width="1.5"/>
  {beams}{labels}
  <line x1="{x0}" y1="{y0 + ph + 10}" x2="{x0 + pw}" y2="{y0 + ph + 10}" stroke="#334155" stroke-width="0.8"/>
  <text x="{x0 + pw/2}" y="{y0 + ph + 22}" text-anchor="middle" font-size="8" fill="#334155">675 cm</text>
  <line x1="{x0 + pw + 10}" y1="{y0}" x2="{x0 + pw + 10}" y2="{y0 + ph}" stroke="#334155" stroke-width="0.8"/>
  <text x="{x0 + pw + 22}" y="{y0 + ph/2}" font-size="8" fill="#334155" transform="rotate(90 {x0 + pw + 22} {y0 + ph/2})">680 cm</text>
  <g transform="translate({pad}, {y0 + ph + 38})">
    <text x="0" y="0" font-size="8" fill="#475569"><tspan font-weight="600">Kot farkları:</tspan> +20,00→+20,20 (0,20) · +20,20→+21,00 (0,80) · +21,00→+21,60 (0,60)</text>
    <text x="0" y="14" font-size="8" fill="#475569">K0601–K0604 (25/60) · PRP.1–4 donatı · GM1 taban (Kesit B-B)</text>
    <text x="0" y="28" font-size="8" fill="#b45309">Ana metraj çatı +16,00 — bu plan üst yapı detayı (+4,00 m fark)</text>
  </g>
</svg>'''


def svg_section_elevation() -> str:
    floors = [
        ("+21,60", "Parapet", 0.6, "#a855f7"),
        ("+21,00", "Duvar/kiriş", 0.8, "#9333ea"),
        ("+20,20", "Döş. üst", 0.2, COLORS["doseme"]),
        ("+20,00", "Kalıp pl.", 0.0, "#c4b5fd"),
        ("+16,00", "Çatı döş.", 0, COLORS["doseme"]),
        ("+12,00", "3. Kat", 3.8, COLORS["perde"]),
        ("+8,00", "2. Kat", 3.8, COLORS["perde"]),
        ("+4,00", "1. Kat", 3.8, COLORS["perde"]),
        ("±0,00", "Zemin", 3.8, COLORS["perde"]),
        ("Bodrum", "H=4,0", 4.0, "#6366f1"),
        ("Radye", "h=0,8", 0.8, COLORS["radye"]),
    ]
    unit = 22
    x0, y0 = 120, 30
    wall_w = 180
    total_h = sum(max(h * unit, 14) for lbl, _, h, _ in floors if h > 0 or lbl in ("+16,00", "+20,00")) + 40

    blocks = []
    y = y0
    for label, sub, h, color in floors:
        if h == 0 and label not in ("+16,00", "+20,00"):
            bh = 14
        else:
            bh = max(h * unit, 14)
        opacity = 0.35 if h > 0 or label in ("+16,00", "+20,00") else 0.15
        blocks.append(
            f'<rect x="{x0}" y="{y}" width="{wall_w}" height="{bh}" fill="{color}" opacity="{opacity}" stroke="{color}" stroke-width="1.5"/>'
            f'<text x="{x0 - 8}" y="{y + bh/2 + 4}" text-anchor="end" font-size="9" font-weight="600" fill="#1e293b">{label}</text>'
            f'<text x="{x0 + wall_w/2}" y="{y + bh/2 + 4}" text-anchor="middle" font-size="8" fill="#475569">{sub}</text>'
        )
        if label in ("±0,00", "+4,00", "+8,00", "+12,00", "+16,00", "+20,00", "Bodrum"):
            blocks.append(
                f'<line x1="{x0 + wall_w + 8}" y1="{y}" x2="{x0 + wall_w + 45}" y2="{y}" stroke="#94a3b8" stroke-dasharray="2,2"/>'
            )
        y += bh

    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 400 {total_h + 50}" width="100%" style="max-width:400px;background:#fafafa">
  <text x="200" y="18" text-anchor="middle" font-size="12" font-weight="700" fill="#1f4e79">04 — KOT KESİTİ (TAM BİNA)</text>
  {''.join(blocks)}
  <text x="200" y="{total_h + 35}" text-anchor="middle" font-size="8" fill="#64748b">Brüt kat 4,00 m · Net perde 3,80 m · Bodrum 4,00 m · +20 üst yapı 1,60 m</text>
</svg>'''


def svg_kirik_formula() -> str:
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 520 220" width="100%" style="max-width:520px;background:#fff">
  <text x="260" y="18" text-anchor="middle" font-size="12" font-weight="700" fill="#1f4e79">05 — KIRIK ÖLÇÜ (a × b × c)</text>
  <g transform="translate(20,35)">
    <rect width="140" height="75" rx="6" fill="#eff6ff" stroke="#2563eb"/>
    <text x="70" y="20" text-anchor="middle" font-size="10" font-weight="600">PERDE</text>
    <text x="70" y="36" text-anchor="middle" font-size="8">a=boy · b=H · c=2</text>
    <text x="70" y="50" text-anchor="middle" font-size="8">Örn: 103,1×3,8×2</text>
    <text x="70" y="64" text-anchor="middle" font-size="8" font-weight="600">= 783,6 m²/kat</text>
  </g>
  <g transform="translate(190,35)">
    <rect width="140" height="75" rx="6" fill="#fef2f2" stroke="#dc2626"/>
    <text x="70" y="20" text-anchor="middle" font-size="10" font-weight="600">KOLON</text>
    <text x="70" y="36" text-anchor="middle" font-size="8">a=2(en+boy) · b=H</text>
    <text x="70" y="50" text-anchor="middle" font-size="8">Örn: 45,96×3,8×13</text>
    <text x="70" y="64" text-anchor="middle" font-size="8" font-weight="600">= 174,6 m²/kat</text>
  </g>
  <g transform="translate(360,35)">
    <rect width="140" height="75" rx="6" fill="#f0fdf4" stroke="#16a34a"/>
    <text x="70" y="20" text-anchor="middle" font-size="10" font-weight="600">DÖŞEME ALT</text>
    <text x="70" y="36" text-anchor="middle" font-size="8">1028,2 − 35,6 − 6,4</text>
    <text x="70" y="50" text-anchor="middle" font-size="8">− 45 boşluk</text>
    <text x="70" y="64" text-anchor="middle" font-size="8" font-weight="600">= 941,3 m²/kat</text>
  </g>
  <g transform="translate(105,125)">
    <rect width="140" height="75" rx="6" fill="#fffbeb" stroke="#d97706"/>
    <text x="70" y="20" text-anchor="middle" font-size="10" font-weight="600">KİRİŞ</text>
    <text x="70" y="36" text-anchor="middle" font-size="8">Yan: ΣL×0,40×2</text>
    <text x="70" y="50" text-anchor="middle" font-size="8">Alt: ΣL×0,325</text>
    <text x="70" y="64" text-anchor="middle" font-size="8" font-weight="600">Toplam 1.441,8 m²</text>
  </g>
  <g transform="translate(275,125)">
    <rect width="140" height="75" rx="6" fill="#f5f3ff" stroke="#7c3aed"/>
    <text x="70" y="20" text-anchor="middle" font-size="10" font-weight="600">RADYE YAN</text>
    <text x="70" y="36" text-anchor="middle" font-size="8">150,5 × 0,80</text>
    <text x="70" y="50" text-anchor="middle" font-size="8">tek yüz çevre</text>
    <text x="70" y="64" text-anchor="middle" font-size="8" font-weight="600">= 120,4 m²</text>
  </g>
  <text x="260" y="215" text-anchor="middle" font-size="9" font-weight="700" fill="#b45309">GENEL TOPLAM: 11.773,4 m² (+ %5 fire → 12.362 m²)</text>
</svg>'''


def svg_doseme_dusum() -> str:
    """Döşeme altı düşüm şeması."""
    items = [
        ("Brüt alan", "51,18 × 20,09", 1028.21, False),
        ("Perde izdüşümü (−)", "CL × t", -35.56, True),
        ("Kolon izdüşümü (−)", "13 ad", -6.36, True),
        ("Merdiven+asansör (−)", "boşluk", -45.00, True),
        ("NET / KAT", "941,29 m²", 941.29, False),
    ]
    y = 55
    bars = ""
    max_w = 280
    for label, sub, val, neg in items:
        w = abs(val) / 1028.21 * max_w
        color = COLORS["doseme"] if not neg else "#ef4444" if neg else COLORS["brand"]
        if label.startswith("NET"):
            color = "#1f4e79"
            bars += (
                f'<rect x="120" y="{y}" width="{w}" height="22" fill="{color}" rx="3"/>'
                f'<text x="115" y="{y + 15}" text-anchor="end" font-size="9" font-weight="700">{label}</text>'
                f'<text x="{120 + w + 6}" y="{y + 15}" font-size="9" font-weight="700">{val:,.2f} m²</text>'
            )
        else:
            bars += (
                f'<rect x="120" y="{y}" width="{w}" height="18" fill="{color}" opacity="{"0.7" if neg else "0.4"}" rx="2"/>'
                f'<text x="115" y="{y + 13}" text-anchor="end" font-size="8">{label}</text>'
                f'<text x="{120 + w + 6}" y="{y + 13}" font-size="8">{sub} · {val:,.2f}</text>'
            )
        y += 28 if label.startswith("NET") else 24

    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 480 {y + 30}" width="100%" style="max-width:480px;background:#fafafa">
  <text x="240" y="22" text-anchor="middle" font-size="12" font-weight="700" fill="#1f4e79">06 — DÖŞEME ALTI DÜŞÜM (±0 ÖRNEK)</text>
  <text x="240" y="38" text-anchor="middle" font-size="9" fill="#64748b">Brüt − perde − kolon − boşluk = net · ×5 kat = 4.706,5 m²</text>
  {bars}
  <text x="240" y="{y + 18}" text-anchor="middle" font-size="8" fill="#64748b">943,4 m²/kat (metraj cetveli yuvarlaması)</text>
</svg>'''


def svg_perde_detay() -> str:
    """Perde kırık ölçü detay kesiti."""
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 420 200" width="100%" style="max-width:420px;background:#fff">
  <text x="210" y="18" text-anchor="middle" font-size="12" font-weight="700" fill="#1f4e79">07 — PERDE KALIP DETAY (KIRIK ÖLÇÜ)</text>
  <!-- duvar kesiti -->
  <rect x="160" y="40" width="30" height="100" fill="#cbd5e1" stroke="#64748b"/>
  <!-- sol kalıp -->
  <rect x="130" y="40" width="30" height="100" fill="{COLORS['perde']}" opacity="0.3" stroke="{COLORS['perde']}"/>
  <!-- sağ kalıp -->
  <rect x="190" y="40" width="30" height="100" fill="{COLORS['perde']}" opacity="0.3" stroke="{COLORS['perde']}"/>
  <!-- ölçü a -->
  <line x1="130" y1="155" x2="220" y2="155" stroke="#334155" marker-end="url(#arr)"/>
  <defs><marker id="arr" markerWidth="6" markerHeight="6" refX="5" refY="3" orient="auto"><path d="M0,0 L6,3 L0,6 Z" fill="#334155"/></marker></defs>
  <text x="175" y="170" text-anchor="middle" font-size="9">a = boy (m)</text>
  <!-- ölçü b -->
  <line x1="235" y1="40" x2="235" y2="140" stroke="#334155"/>
  <text x="250" y="95" font-size="9">b = H (m)</text>
  <text x="250" y="108" font-size="8" fill="#64748b">3,80 üst kat</text>
  <text x="250" y="120" font-size="8" fill="#64748b">4,00 bodrum</text>
  <!-- c -->
  <text x="175" y="35" text-anchor="middle" font-size="9" font-weight="600">c = 2 (çift yüz)</text>
  <text x="210" y="195" text-anchor="middle" font-size="9" fill="#2563eb" font-weight="600">m² = a × b × 2 − boşluk düşümü</text>
</svg>'''


def write_svgs():
    CIZIM.mkdir(parents=True, exist_ok=True)
    files = {
        "01_tipik_kat_plani.svg": svg_plan_typical(),
        "02_temel_plan_cad.svg": svg_temel_plan(),
        "03_cati_20_plan.svg": svg_cati_plan(),
        "04_kot_kesit.svg": svg_section_elevation(),
        "05_kirik_olcu_sema.svg": svg_kirik_formula(),
        "06_doseme_dusum.svg": svg_doseme_dusum(),
        "07_perde_detay.svg": svg_perde_detay(),
    }
    for name, content in files.items():
        (CIZIM / name).write_text(content, encoding="utf-8")
    return files


def write_html(svg_files: dict):
    titles = {
        "01_tipik_kat_plani.svg": ("Tipik Kat Planı", "Üst katlar — perde, kolon, döşeme, kiriş"),
        "02_temel_plan_cad.svg": ("Temel / Bodrum Planı", "CAD eleman numaraları — P109, BP1, BOŞLUK…"),
        "03_cati_20_plan.svg": ("+20,00 Kalıp Planı", "Çatı üstü yapı — D601/D602, K0601–K0604"),
        "04_kot_kesit.svg": ("Kot Kesiti", "Radye −5,90 → +21,60 tam merdiven"),
        "05_kirik_olcu_sema.svg": ("Kırık Ölçü Formülleri", "a×b×c — poz bazlı örnekler"),
        "06_doseme_dusum.svg": ("Döşeme Altı Düşüm", "Brüt 1.028 → net 941 m²/kat"),
        "07_perde_detay.svg": ("Perde Kalıp Detay", "Çift yüz kırık ölçü kesiti"),
    }
    cards = ""
    for fname, svg in svg_files.items():
        title, sub = titles[fname]
        cards += f'''
<section class="card" id="{fname.replace(".svg","")}">
  <h2>{title}</h2>
  <p class="sub">{sub}</p>
  <div class="svg-wrap">{svg}</div>
</section>'''

    kot_rows = ""
    for kat, sub, h, top, d in KOTLAR:
        kot_rows += (
            f"<tr><td>{kat} ({sub})</td><td>{h}</td>"
            f"<td>{d['perde'] or '—'}</td><td>{d['kolon'] or '—'}</td>"
            f"<td>{d['doseme'] or '—'}</td><td>{d['kiris'] or '—'}</td><td><b>{top}</b></td></tr>"
        )

    poz_rows = ""
    for kod, tarif, m2, formul, pay in POZlar:
        poz_rows += f"<tr><td>{kod}</td><td>{tarif}</td><td class='num'>{m2:,.1f}</td><td>{formul}</td><td class='num'>{pay}</td></tr>"

    html = f'''<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover"/>
<meta name="theme-color" content="#1f4e79"/>
<title>Kalıp Metraj Detay Çizim</title>
<style>
:root{{--brand:#1f4e79;--card:#fff;--muted:#64748b;--line:#e2e8f0;--accent:#f5a623}}
*{{box-sizing:border-box}}body{{margin:0;font-family:system-ui,sans-serif;background:#eef2f7;color:#1a1a1a;line-height:1.45}}
header{{background:linear-gradient(160deg,#1f4e79,#0f2744);color:#fff;padding:1.2rem 1rem;text-align:center}}
header h1{{margin:0;font-size:1.15rem}}header p{{margin:.35rem 0 0;font-size:.82rem;opacity:.9}}
main{{max-width:580px;margin:0 auto;padding:1rem 1rem 2rem}}
.card{{background:var(--card);border-radius:12px;padding:1rem;margin:.75rem 0;box-shadow:0 4px 16px rgba(0,0,0,.06)}}
.card h2{{margin:0;font-size:1rem;color:var(--brand)}}
.sub{{margin:.25rem 0 .75rem;font-size:.8rem;color:var(--muted)}}
.svg-wrap{{overflow-x:auto;border:1px solid var(--line);border-radius:8px;background:#fafafa}}
table{{width:100%;border-collapse:collapse;font-size:.78rem}}
td,th{{padding:.35rem .2rem;border-bottom:1px solid var(--line);text-align:right;vertical-align:top}}
th{{text-align:left;color:var(--muted)}}
td.num{{font-variant-numeric:tabular-nums}}
.scroll-x{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
.back{{display:block;padding:1rem;color:var(--brand);font-weight:600;text-decoration:none}}
.nav{{display:flex;flex-wrap:wrap;gap:.35rem;margin:.5rem 0}}
.nav a{{font-size:.68rem;padding:.3rem .5rem;background:#fff;border:1px solid var(--line);border-radius:6px;color:var(--brand);text-decoration:none;font-weight:600}}
.hero-stat{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:.5rem;margin:.75rem 0}}
.hero-stat div{{background:linear-gradient(135deg,#1e3a5f,#2563eb);color:#fff;border-radius:10px;padding:.65rem;text-align:center;font-size:.72rem}}
.hero-stat b{{display:block;font-size:1rem;color:var(--accent)}}
</style>
</head>
<body>
<a class="back" href="index.html">← Portal</a>
<header>
  <h1>Kalıp Metraj Detay Çizimi</h1>
  <p>Karşıyaka Ortaokulu · 11.773,4 m² · 7 çizim + tablolar</p>
</header>
<main>
<div class="hero-stat">
  <div><b>11.773 m²</b>Toplam metraj</div>
  <div><b>7</b>Detay çizim</div>
  <div><b>450–500</b>TL/m² BF</div>
</div>
<nav class="nav">
  <a href="#01_tipik_kat_plani">Tipik kat</a>
  <a href="#02_temel_plan_cad">Temel</a>
  <a href="#03_cati_20_plan">+20 çatı</a>
  <a href="#04_kot_kesit">Kesit</a>
  <a href="#05_kirik_olcu_sema">Kırık ölçü</a>
  <a href="#06_doseme_dusum">Düşüm</a>
  <a href="#07_perde_detay">Perde</a>
</nav>
{cards}
<div class="card scroll-x">
  <h2>Poz Bazlı Metraj (m²)</h2>
  <p class="sub">15.180.1002 poz grupları — kırık ölçü cetveli</p>
  <table>
    <tr><th>Poz</th><th>Tarif</th><th class="num">m²</th><th>Formül</th><th class="num">Pay</th></tr>
    {poz_rows}
    <tr><th colspan="2">GENEL TOPLAM</th><th class="num"><b>11.773,4</b></th><th colspan="2">+%5 fire → 12.362</th></tr>
  </table>
</div>
<div class="card scroll-x">
  <h2>Kat Bazlı Metraj (m²)</h2>
  <table>
    <tr><th>Kot/Kat</th><th>H(m)</th><th>Perde</th><th>Kolon</th><th>Döşeme</th><th>Kiriş</th><th>Toplam</th></tr>
    {kot_rows}
    <tr><th colspan="6">GENEL TOPLAM</th><th><b>11.773,4</b></th></tr>
  </table>
</div>
<div class="card">
  <h2>İlgili sayfalar</h2>
  <p class="sub"><a href="kot-farklari.html">Kot farkları tablosu</a> · <a href="telefon.html">Özet tablo</a> · <a href="metraj.html">Metraj tabloları</a></p>
</div>
</main>
</body>
</html>'''
    OUT_HTML.write_text(html, encoding="utf-8")


def write_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Metraj Detay"
    hdr = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="FFAAAAAA")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = "KALIP METRAJ DETAY ÇİZİM CETVELİ — Karşıyaka Ortaokulu"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    headers = ["Sıra", "Bölüm", "Kot", "Net H", "Eleman", "a (m)", "b (m)", "c", "m²", "Formül / Not"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = hdr
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = brd

    rows = [
        (1, "Radye", "Temel", "0,80", "Yan kenar 1-12", "150,5", "0,80", "1", 120.4, "Çevre × h"),
        (2, "Bodrum", "±0 alt", "4,00", "Perde çift yüz", "133,8", "4,00", "2", 1152.7, "Boşluk −19,2"),
        (3, "Bodrum", "±0 alt", "4,00", "Kolon", "45,96", "4,00", "13", 184.0, "13 adet"),
        (4, "Üst", "0/+4/+8/+12", "3,80", "Perde ×4", "103,1", "3,80", "2×4", 3468.8, "Boşluk −51,2"),
        (5, "Üst", "0/+4/+8/+12", "3,80", "Kolon ×4", "45,96", "3,80", "13×4", 699.2, ""),
        (6, "Döşeme", "5 kat", "—", "Alt kalıp", "941,29", "1", "5", 4706.5, "Brüt−düşüm"),
        (7, "Kiriş", "5 kat", "0,40/0,325", "Yan+alt", "ΣL", "—", "—", 1441.8, "H1 20/60 vb."),
    ]
    for i, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(row=i, column=c, value=val).border = brd

    r = 4 + len(rows) + 1
    ws.cell(row=r, column=8, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=r, column=9, value=11773.4).font = Font(bold=True)

    ws2 = wb.create_sheet("Kat Kırılım")
    h2 = ["Kat", "Perde", "Kolon", "Döşeme", "Kiriş", "Toplam"]
    for c, h in enumerate(h2, 1):
        ws2.cell(row=1, column=c, value=h).font = Font(bold=True)
    for i, (kat, _, _, top, d) in enumerate(KOTLAR, 2):
        ws2.cell(row=i, column=1, value=kat)
        ws2.cell(row=i, column=2, value=d["perde"] or "")
        ws2.cell(row=i, column=3, value=d["kolon"] or "")
        ws2.cell(row=i, column=4, value=d["doseme"] or "")
        ws2.cell(row=i, column=5, value=d["kiris"] or "")
        ws2.cell(row=i, column=6, value=top)

    ws3 = wb.create_sheet("CAD Elemanlar")
    cad = [
        ("Temel", "P109", "30/160", "Perde", "Bodrum metraj"),
        ("Temel", "P103", "280/35", "Perde", ""),
        ("Temel", "BP1", "35/575", "Bodrum perde", ""),
        ("Temel", "BOŞLUK", "100/220", "Düşüm", "Metrajdan düşülür"),
        ("Temel", "RD01", "h=80", "Radye", "T.Ü.K. −5,10"),
        ("Çatı +20", "D601/D602", "h=20", "Döşeme", "650×650 iç"),
        ("Çatı +20", "K0601–K0604", "25/60", "Kiriş", ""),
    ]
    for c, h in enumerate(["Plan", "Etiket", "Boyut", "Tip", "Not"], 1):
        ws3.cell(row=1, column=c, value=h).font = Font(bold=True)
    for i, row in enumerate(cad, 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=i, column=c, value=val)

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(OUT_XLSX)


def main():
    svg_files = write_svgs()
    write_html(svg_files)
    write_excel()
    print(f"SVG: {len(svg_files)} files → {CIZIM}")
    print(f"HTML: {OUT_HTML}")
    print(f"Excel: {OUT_XLSX}")


if __name__ == "__main__":
    main()
